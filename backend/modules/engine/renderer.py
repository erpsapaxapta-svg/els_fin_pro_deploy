from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _style_header(ws, row: int, last_col: int):
    fill = PatternFill("solid", fgColor="F2F2F2")
    font = Font(bold=True)
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_caption(ws, row: int, last_col: int):
    fill = PatternFill("solid", fgColor="FFF5D6")
    font = Font(bold=True)
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font


def _style_total(ws, row: int, last_col: int):
    font = Font(bold=True)
    top = Border(top=Side(style="thin"))
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = top


def _apply_table_borders(ws, start_row: int, end_row: int, last_col: int):
    thin = Side(style="thin")
    for r in range(start_row, end_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).border = Border(
                left=thin if c == 1 else None,
                right=thin if c == last_col else None,
                top=thin if r == start_row else None,
                bottom=thin if r == end_row else None,
            )


def _autosize(ws, last_col: int, min_w: int = 12, max_w: int = 28):
    for c in range(1, last_col + 1):
        col = get_column_letter(c)
        maxlen = 0
        for cell in ws[col]:
            try:
                maxlen = max(maxlen, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col].width = max(min_w, min(max_w, maxlen + 2))


def render_table(*, template: Dict, periods: List[str], company: Optional[str], out_path: Path) -> Path:
    """
    ÙŠØ¨Ù†ÙŠ Ø´ÙŠØª Excel Ø­Ø³Ø¨ Ù‚Ø§Ù„Ø¨ YAML.
    - Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© = ["Description"] + periods
    - rows: Ø¹Ù†Ø§ØµØ± Ù…Ù† Ø«Ù„Ø§Ø« Ø£Ù†ÙˆØ§Ø¹: caption | item | total
      - caption: Ø¹Ù†ÙˆØ§Ù† Ù‚Ø³Ù…
      - item: Ø¨Ù†Ø¯ Ø¹Ø§Ø¯ÙŠ (Ù‚ÙŠÙ…Ù‡ Ø­Ø§Ù„ÙŠØ§Ù‹ 0 â€” Ù†Ø±Ø¨Ø·Ù‡Ø§ Ø¨Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù„Ø§Ø­Ù‚Ù‹Ø§)
      - total: ÙŠØ¬Ù…Ø¹:
          1) Ù…Ù† Ø£ÙˆÙ„ Ø¨Ù†Ø¯ Ø¨Ø¹Ø¯ Ø¢Ø®Ø± caption Ø­ØªÙ‰ Ø§Ù„Ø³Ø·Ø± Ø§Ù„Ø³Ø§Ø¨Ù‚ (Ù„Ùˆ Ù…ÙÙŠØ´ sum_labels)
          2) Ø£Ùˆ ÙŠØ¬Ù…Ø¹ Ù…Ø¬Ù…ÙˆØ¹Ø© totals Ø³Ø§Ø¨Ù‚Ø© Ø¨Ø§Ù„Ø§Ø³Ù… (sum_labels: ["Total X", "Total Y"])
    """
    sheet_name = template.get("sheet_name", "Sheet1")
    title = template.get("title", sheet_name)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Ø¹Ù†ÙˆØ§Ù†
    ws.cell(row=1, column=1, value=title + (f" â€” {company}" if company else ""))
    ws["A1"].font = Font(bold=True, size=14)

    # Ù‡ÙŠØ¯Ø±
    header_row = 4
    headers = ["Description"] + periods
    last_col = len(headers)
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=h)
    _style_header(ws, header_row, last_col)

    # Ø¬Ø³Ù… Ø§Ù„Ø¬Ø¯ÙˆÙ„
    current_row = header_row + 1
    section_start_row = None  # Ø£ÙˆÙ„ Ø¨Ù†Ø¯ Ø¨Ø¹Ø¯ Ø§Ù„ÙƒØ§Ø¨ØªØ´Ù†
    totals_rows: Dict[str, int] = {}  # Ø§Ø³Ù… Ø§Ù„ØªÙˆØªØ§Ù„ -> Ø±Ù‚Ù… Ø§Ù„ØµÙ

    for spec in template.get("rows", []):
        if "caption" in spec:
            ws.cell(row=current_row, column=1, value=spec["caption"])
            _style_caption(ws, current_row, last_col)
            section_start_row = current_row + 1
            current_row += 1

        elif "item" in spec:
            ws.cell(row=current_row, column=1, value=spec["item"])
            # Ø§Ù„Ù‚ÙŠÙ… placeholder = 0 (Ù‡Ù†Ø±Ø¨Ø·Ù‡Ø§ Ø¨Ø¨ÙŠØ§Ù†Ø§Øª ÙØ¹Ù„ÙŠØ© Ù„Ø§Ø­Ù‚Ù‹Ø§)
            for j in range(2, last_col + 1):
                ws.cell(row=current_row, column=j, value=0).number_format = "#,##0"
            current_row += 1

        elif "total" in spec:
            label = spec["total"]
            ws.cell(row=current_row, column=1, value=label)
            # ØµÙŠØºØ© Ø§Ù„Ø¬Ù…Ø¹
            for j in range(2, last_col + 1):
                col_letter = get_column_letter(j)
                # Ø­Ø§Ù„Ø© sum_labels (ÙŠØ¬Ù…Ø¹ totals Ø³Ø§Ø¨Ù‚Ø© Ø¨Ø£Ø³Ù…Ø§Ø¦Ù‡Ø§)
                if "sum_labels" in spec:
                    coords = []
                    for nm in spec["sum_labels"]:
                        r = totals_rows.get(nm)
                        if r:
                            coords.append(f"{col_letter}{r}")
                    formula = f"={' + '.join(coords)}" if coords else "0"
                else:
                    start = section_start_row or current_row
                    end = current_row - 1
                    if end >= start:
                        formula = f"=SUM({col_letter}{start}:{col_letter}{end})"
                    else:
                        formula = "0"
                ws.cell(row=current_row, column=j, value=formula).number_format = "#,##0"
            _style_total(ws, current_row, last_col)
            totals_rows[label] = current_row
            current_row += 1

    # Ø­Ø¯ÙˆØ¯ Ø­ÙˆÙ„ Ø§Ù„Ø¬Ø¯ÙˆÙ„
    _apply_table_borders(ws, header_row, current_row - 1, last_col)

    # ØªÙ†Ø³ÙŠÙ‚ Ø¹Ø±Ø¶ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
    _autosize(ws, last_col)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path

