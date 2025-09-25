# modules/filters_and_style.py
from pathlib import Path
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import shutil
import tempfile

# ====== Ø£Ù„ÙˆØ§Ù† ÙˆÙ‡ÙˆÙŠØ© Ø«Ø§Ø¨ØªØ© ======
GREEN_DARK  = "2F6E3A"   # Header (ØºØ§Ù…Ù‚)
GREEN_LIGHT = "CFE6CF"   # Ø´Ø±ÙŠØ· ÙØ§ØªØ­ Ø£Ø¹Ù„Ù‰
GREEN_MED   = "6AA958"   # Ø¹Ù†Ø§ÙˆÙŠÙ† ÙØªØ±Ø§Øª
TOTAL_FILL  = "F7E3B0"   # ØµÙÙˆÙ Total (ÙØ§ØªØ­)
BORDER_COL  = "888888"

def normalize_periods(selected: List[str], canonical: List[str]) -> List[str]:
    """ÙŠØ­Ø§ÙØ¸ Ø¹Ù„Ù‰ ØªØ±ØªÙŠØ¨ periods ØªØ¨Ø¹ canonical ÙˆÙŠØ´ÙŠÙ„ Ø§Ù„Ø¯ÙˆØ¨Ù„ÙƒÙŠØª"""
    want = [str(p).upper() for p in (selected or [])]
    seen = set()
    out = []
    for p in canonical:
        u = str(p).upper()
        if u in want and u not in seen:
            out.append(p)
            seen.add(u)
    return out

def filter_input_excel(input_path: Path, company: Optional[str], periods: Optional[List[str]]) -> Path:
    """
    Ø­Ø§Ù„ÙŠØ§Ù‹ placeholder: Ù…Ø¬Ø±Ø¯ Ù†Ø³Ø® Ø§Ù„Ù…Ù„Ù Ø¥Ù„Ù‰ temp.
    Ù„Ø§Ø­Ù‚Ø§Ù‹ ØªÙ‚Ø¯Ø± ØªØ¹Ù…Ù„ ÙÙ„ØªØ±Ø© ÙØ¹Ù„ÙŠØ© Ù„Ù„Ù€ RawTB Ù„Ùˆ Ø­Ø¨ÙŠØª.
    """
    tempdir = Path(tempfile.mkdtemp())
    out = tempdir / input_path.name
    shutil.copy2(str(input_path), str(out))
    return out

def _auto_fit(ws, start_col=1, end_col=None, minw=9, maxw=32):
    end_col = end_col or ws.max_column
    widths = {}
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        widths[letter] = minw
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=min(400, ws.max_row)):
            val = row[0].value
            if val is None: continue
            l = len(str(val))
            if l > widths[letter]:
                widths[letter] = min(maxw, l + 2)
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

def pretty_format_workbook(xlsx_path: Path):
    """ØªØ·Ø¨ÙŠÙ‚ Ø´ÙƒÙ„ Ø«Ø§Ø¨Øª: Ø±Ø£Ø³ Ø£Ø®Ø¶Ø±ØŒ ØªØ¬Ù…ÙŠØ¯ C3ØŒ Totals Ø¨Ù„ÙˆÙ† ÙØ§ØªØ­ØŒ Ø£Ø±Ù‚Ø§Ù… Ø³Ø§Ù„Ø¨Ø© Ø­Ù…Ø±Ø§Ø¡ØŒ Ø­Ø¯ÙˆØ¯ Ø®ÙÙŠÙØ©."""
    wb = load_workbook(filename=str(xlsx_path))
    thin = Side(style="thin", color=BORDER_COL)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    number_style = NamedStyle(name="els_number")
    number_style.number_format = '#,##0;[Red]-#,##0'
    number_style.border = border
    number_style.alignment = Alignment(horizontal="right", vertical="center")

    for ws in wb.worksheets:
        # Freeze: Ø¹Ù…ÙˆØ¯ÙŠÙ† + ØµÙÙŠÙ†
        ws.freeze_panes = "C3"

        # Ø£ÙˆÙ„ ØµÙ: Ø´Ø±ÙŠØ· Ø£Ø®Ø¶Ø± ÙØ§ØªØ­
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)

        # ØµÙ 2: Ø¹Ù†Ø§ÙˆÙŠÙ† ÙØªØ±Ø§Øª/Ø¹Ù†Ø§ÙˆÙŠÙ† Ø±Ø¦ÙŠØ³ÙŠØ©
        for c in ws[2]:
            c.fill = PatternFill("solid", fgColor=GREEN_MED)
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

        # Ø¹Ù…ÙˆØ¯ B (Ø£Ø³Ù…Ø§Ø¡ Ø§Ù„Ø¨Ù†ÙˆØ¯) ØªÙ†Ø³ÙŠÙ‚ Ù†Øµ ØºØ§Ù…Ù‚ Ø®ÙÙŠÙ
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=2)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            txt = str(cell.value or "")
            if txt.strip().lower().startswith("total"):
                for c in ws[r]:
                    c.fill = PatternFill("solid", fgColor=TOTAL_FILL)
                    c.font = Font(bold=True)
                    c.border = border

        # Ø£Ø±Ù‚Ø§Ù…: Ù…Ù† Ø§Ù„Ø¹Ù…ÙˆØ¯ 3 Ù„Ù„Ù†Ù‡Ø§ÙŠØ©
        for r in range(3, ws.max_row + 1):
            for c in range(3, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.style = number_style
                cell.border = border

        # ØµÙ Ø§Ù„Ø¹Ù†Ø§ÙˆÙŠÙ† Ø§Ù„Ø¹Ù„ÙˆÙ‰ Ø¬Ø¯Ø§Ù‹ (Ù„Ùˆ ÙˆØ¬Ø¯)
        # Ù…Ù…ÙƒÙ† ÙŠÙƒÙˆÙ† Ù…ØªØ±ÙˆÙƒ ÙØ§Ø±Øº â€” Ù†Ø¶Ø¹ Ù„ÙˆÙ† Ø«Ø§Ø¨Øª Ø§Ù„ØºØ§Ù…Ù‚
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
            c.border = border

        # ØªØ­Ø³ÙŠÙ† Ø¹Ø±Ø¶ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
        _auto_fit(ws, start_col=1, end_col=ws.max_column)

    if "els_number" not in wb.named_styles:
        wb.add_named_style(number_style)

    wb.save(str(xlsx_path))

